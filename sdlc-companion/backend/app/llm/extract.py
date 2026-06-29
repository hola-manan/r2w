"""Document text extraction for user attachments (docx / xlsx).

Users hand the agent an existing brief — a Word doc ~90% of the time, a
spreadsheet occasionally — to *show intent* instead of retyping it. We parse the
file to plain text server-side and fold that text into the conversation message;
Gemini does not ingest docx/xlsx natively, so text extraction is the right path.

Only `.docx` and `.xlsx` are accepted; anything else raises `ValueError`.
"""
from __future__ import annotations

import io

SUPPORTED_EXTENSIONS = (".docx", ".xlsx")


def extract_docx(data: bytes) -> str:
    """Return the raw text of a .docx file."""
    import mammoth

    result = mammoth.extract_raw_text(io.BytesIO(data))
    return (result.value or "").strip()


def extract_xlsx(data: bytes) -> str:
    """Return a tab-separated text rendering of a .xlsx workbook.

    Each sheet is prefixed with a `# <sheet name>` header; rows become
    tab-joined lines with trailing empty cells trimmed. Fully empty rows are
    skipped so the model sees only meaningful content.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        blocks: list[str] = []
        for ws in wb.worksheets:
            lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    lines.append("\t".join(cells))
            if lines:
                blocks.append(f"# {ws.title}\n" + "\n".join(lines))
        return "\n\n".join(blocks).strip()
    finally:
        wb.close()


def extract(filename: str, data: bytes) -> str:
    """Dispatch to the right parser by file extension.

    Raises `ValueError` for unsupported extensions or unparseable content.
    """
    name = (filename or "").lower()
    if name.endswith(".docx"):
        return extract_docx(data)
    if name.endswith(".xlsx"):
        return extract_xlsx(data)
    raise ValueError(
        f"unsupported file type for '{filename}'; only .docx and .xlsx are accepted"
    )
